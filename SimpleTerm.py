import openpyxl
import tkinter as tk
from tkinter import messagebox, filedialog
import pyperclip
import os

class SimpleTerm:
    def __init__(self, root, file_path=None):
        self.root = root
        self.file_path = file_path
        self.termbase = {}
        self.results = []
        self.current_index = 0

        if self.file_path is None or not self.file_path.endswith('.xlsx'):
            self.select_excel_file()
        else:
            self.load_excel()
            self.setup_gui()

    def select_excel_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", ".xlsx")],
            title="Select an Excel file as the termbase"
        )
        if file_path:
            self.file_path = file_path
            self.load_excel()
            if self.termbase:
                self.setup_gui()
        else:
            messagebox.showwarning("No File Selected", "No File Selected. Shutting down")
            exit()

    def load_excel(self):
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                source_term, target_term, notes = row
                if source_term:
                    if source_term.lower() not in self.termbase:
                        self.termbase[source_term.lower()] = []
                    self.termbase[source_term.lower()].append({'target_term': target_term, 'notes': notes})
        except Exception as e:
            messagebox.showerror("Error", f"Error loading Excel file:\n{str(e)}")

    def find_equivalent(self, term):
        term = term.lower()
        return self.termbase.get(term, [])

    def search_term(self, event=None):
        term = self.entry.get().strip()

        if term and self.termbase:
            try:
                self.results = self.find_equivalent(term)
                self.current_index = 0
                self.update_display()
            except Exception as e:
                print(f"Error occurred: {str(e)}")  # Debug print
                messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        else:
            messagebox.showwarning("Input Error", "Please enter a term to search or load the Excel file first.")

    def update_display(self):
        if self.results:
            result = self.results[self.current_index]
            self.result_label.config(text=result['target_term'])

            notes_text = result['notes'] if result['notes'] else ""
            self.notes_text.config(state=tk.NORMAL)
            self.notes_text.delete(1.0, tk.END)
            self.notes_text.insert(tk.END, notes_text)
            self.notes_text.config(state=tk.DISABLED)

            self.result_label.config(fg='red' if len(self.results) > 1 else 'black')
        else:
            self.result_label.config(text='Term not found. Do you want to add a new term?\n(ctrl+n)')
            self.notes_text.config(state=tk.NORMAL)
            self.notes_text.delete(1.0, tk.END)
            self.notes_text.config(state=tk.DISABLED)

        self.root.update_idletasks()
        self.root.geometry(f"{self.root.winfo_width()}x{self.root.winfo_height()}")

    def setup_gui(self):
        if self.file_path:
            file_name = os.path.basename(self.file_path)
            self.root.title(f"SimpleTerm: {file_name}")

        self.root.attributes('-topmost', True)
        self.root.configure(bg='#f0f0f0')

        input_frame = tk.Frame(self.root, padx=10, pady=10, bg='#f0f0f0')
        input_frame.pack(padx=10, pady=10, fill=tk.X)
        input_frame.pack_propagate(False)

        self.entry = tk.Entry(input_frame, font=('Arial', 14), relief=tk.FLAT, width=20)
        self.entry.grid(row=0, column=0, padx=(0, 10), sticky='ew')

        self.result_label = tk.Label(input_frame, text='', font=('Arial', 14), anchor='w', bg='#f0f0f0')
        self.result_label.grid(row=0, column=1, padx=(10, 0), sticky='ew')

        input_frame.grid_columnconfigure(0, weight=0)

        self.entry.bind('<Return>', self.search_term)

        result_frame = tk.Frame(self.root, padx=10, pady=0, bg='#f0f0f0')
        result_frame.pack(padx=10, pady=0, fill=tk.BOTH, expand=True)
        self.notes_text = tk.Text(result_frame, wrap=tk.WORD, font=('Arial', 12), state=tk.DISABLED, bg='#f0f0f0', relief=tk.FLAT)
        self.notes_text.pack(fill=tk.BOTH, expand=True)
        result_frame.pack_propagate(False)

        self.root.bind('<Right>', self.navigate_results)
        self.root.bind('<Left>', self.navigate_results)
        self.root.bind('<Tab>', self.navigate_results)
        self.root.bind('<Control-c>', self.copy_result_term)
        self.root.bind('<Control-n>', self.add_new_term)
        self.root.bind('<Control-plus>', self.increase_font_size)
        self.root.bind('<Control-minus>', self.decrease_font_size)
        self.root.bind('<Control-Shift-plus>', self.increase_notes_font_size)
        self.root.bind('<Control-Shift-minus>', self.decrease_notes_font_size)
        self.root.bind('<F5>', self.refresh_excel)
        self.root.bind('<F3>', self.display_help)
        self.root.bind('<F2>', self.change_excel_file)

    def navigate_results(self, event):
        if self.results:
            if event.keysym == 'Right' or event.keysym == 'Tab':
                self.current_index = (self.current_index + 1) % len(self.results)
            elif event.keysym == 'Left':
                self.current_index = (self.current_index - 1) % len(self.results)
            self.update_display()

    def copy_result_term(self, event):
        if self.results:
            pyperclip.copy(self.results[self.current_index]['target_term'])
            self.result_label.config(bg='green')
            self.root.after(500, lambda: self.result_label.config(bg=self.root.cget('bg')))

    def increase_font_size(self, event=None):
        current_font_size = int(self.entry.cget('font').split()[1])
        new_font_size = current_font_size + 1
        self.entry.config(font=('Arial', new_font_size))
        self.result_label.config(font=('Arial', new_font_size))

    def decrease_font_size(self, event=None):
        current_font_size = int(self.entry.cget('font').split()[1])
        new_font_size = max(current_font_size - 1, 8)
        self.entry.config(font=('Arial', new_font_size))
        self.result_label.config(font=('Arial', new_font_size))

    def increase_notes_font_size(self, event=None):
        current_font_size = int(self.notes_text.cget('font').split()[1])
        new_font_size = current_font_size + 1
        self.notes_text.config(font=('Arial', new_font_size))

    def decrease_notes_font_size(self, event=None):
        current_font_size = int(self.notes_text.cget('font').split()[1])
        new_font_size = max(current_font_size - 1, 8)
        self.notes_text.config(font=('Arial', new_font_size))

    def refresh_excel(self, event=None):
        self.load_excel()
        if self.termbase:
            file_name = os.path.basename(self.file_path) if self.file_path else ""
            self.root.title(f"SimpleTerm: {file_name}")
            self.root.after(2000, lambda: self.root.title(f"SimpleTerm: {file_name}"))
            self.root.after(0, lambda: self.root.title("Updated!"))  # Display "Updated!" briefly

    def change_excel_file(self, event=None):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", ".xlsx")],
            title="Select a new Excel file"
        )
        if file_path:
            self.file_path = file_path
            self.load_excel()
            if self.termbase:
                file_name = os.path.basename(self.file_path)
                self.root.title(f"SimpleTerm: {file_name}")
                self.results = []
                self.update_display()
                self.root.after(2000, lambda: self.root.title(f"SimpleTerm: {file_name}"))
                self.root.after(0, lambda: self.root.title("File Changed!"))  # Display "File Changed!" briefly

    def display_help(self, event=None):
        help_text = (
            "The first row of the Excel file should be 'Source Term', 'Target Term', 'Notes'\n"
            "Shortcuts\n"
            "Tab Key, Left and Right arrows: Navigate between results\n"
            "Ctrl+C: Copy the displayed result\n"
            "Ctrl++/-: Increase or decrease the font size\n"
            "Ctrl+Shift++/-: Increase or decrease the font size of notes\n"
            "F2: Change the Excel file used\n"
            "F3: Open this help message\n"
            "F5: Refresh the Excel database\n"
            "Ctrl+N: Add a new term"
        )
        messagebox.showinfo("Help", help_text)

    def add_new_term(self, event=None):
        add_window = tk.Toplevel(self.root)
        add_window.title("Add New Term")
        add_window.configure(bg='#f0f0f0')
        
        # Ensure the add_window is on top and focused
        add_window.focus_force()
        
        tk.Label(add_window, text="Source Term:", font=('Arial', 12), bg='#f0f0f0').grid(row=0, column=0, padx=10, pady=5)
        source_entry = tk.Entry(add_window, font=('Arial', 12), relief=tk.FLAT)
        source_entry.grid(row=0, column=1, padx=10, pady=5)
        
        # Set focus to Source Term entry
        source_entry.focus_set()

        term_to_add = self.entry.get().strip()
        if term_to_add:
            source_entry.insert(tk.END, term_to_add)

        tk.Label(add_window, text="Target Term:", font=('Arial', 12), bg='#f0f0f0').grid(row=1, column=0, padx=10, pady=5)
        target_entry = tk.Entry(add_window, font=('Arial', 12), relief=tk.FLAT)
        target_entry.grid(row=1, column=1, padx=10, pady=5)

        tk.Label(add_window, text="Notes:", font=('Arial', 12), bg='#f0f0f0').grid(row=2, column=0, padx=10, pady=5)
        notes_entry = tk.Entry(add_window, font=('Arial', 12), relief=tk.FLAT)
        notes_entry.grid(row=2, column=1, padx=10, pady=5)

        def save_new_term(event=None):
            source_term = source_entry.get().strip()
            target_term = target_entry.get().strip()
            notes = notes_entry.get().strip()

            if source_term and target_term:
                if source_term.lower() not in self.termbase:
                    self.termbase[source_term.lower()] = []
                self.termbase[source_term.lower()].append({'target_term': target_term, 'notes': notes})

                try:
                    workbook = openpyxl.load_workbook(self.file_path)
                    sheet = workbook.active
                    sheet.append([source_term, target_term, notes])
                    workbook.save(self.file_path)
                    messagebox.showinfo("Success", "New term added successfully!")
                    add_window.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Error saving to Excel file:\n{str(e)}")
            else:
                messagebox.showwarning("Input Error", "Please enter both Source Term and Target Term.")

        save_button = tk.Button(add_window, text="Save", font=('Arial', 12), relief=tk.FLAT, command=save_new_term)
        save_button.grid(row=3, columnspan=2, pady=10)

        # Bind Enter key to save_new_term function
        add_window.bind('<Return>', save_new_term)

if __name__ == "__main__":
    root = tk.Tk()
    app = SimpleTerm(root)
    root.mainloop()
